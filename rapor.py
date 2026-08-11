"""
Tarama sonucunu Excel (.xlsx) kaydı olarak üretir.

Babanın iş akışı: gün sonunda tarar → ertesi gün açılışında takip eder.
Bu yüzden kayıtta net olarak: TARAMA TARİHİ+SAATİ ve her hissenin O ANKİ FİYATI.
Değişim yüzdesi yeşil (+) / kırmızı (−) renklenir.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

_KOYU = "FF1F2328"
_GRI = "FF6E7681"
_YESIL = "FF1A7F37"
_KIRMIZI = "FFB42318"
_BASLIK_BG = "FF2F81F7"
_BEYAZ = "FFFFFFFF"
_ince = Side(style="thin", color="FFD0D7DE")


def tarama_excel(satirlar, tarama_zamani, veri_saati=None) -> bytes:
    """
    satirlar: [{hisse, fiyat(float|None), degisim(float|None), rating, sektor}]
    tarama_zamani: "11.06.2026 18:45" (taramanın yapıldığı tam an)
    veri_saati: opsiyonel "11.06 18:09" (TradingView veri saati)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarama"

    ws["A1"] = "BIST Tarama Kaydı"
    ws["A1"].font = Font(bold=True, size=14, color=_KOYU)
    ws["A2"] = f"Tarama zamanı: {tarama_zamani}"
    ws["A2"].font = Font(bold=True, size=11, color=_KOYU)
    if veri_saati:
        ws["A3"] = f"Veri saati: {veri_saati}  (15 dk gecikmeli)"
        ws["A3"].font = Font(size=10, color=_GRI)

    bas = 5
    basliklar = ["Hisse", "Fiyat (₺)", "Değişim %", "Rating", "Sektör"]
    for j, b in enumerate(basliklar, start=1):
        c = ws.cell(row=bas, column=j, value=b)
        c.font = Font(bold=True, color=_BEYAZ)
        c.fill = PatternFill("solid", fgColor=_BASLIK_BG)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=_ince)

    for i, s in enumerate(satirlar, start=bas + 1):
        ws.cell(row=i, column=1, value=s.get("hisse", "")).font = Font(bold=True, color=_KOYU)

        fc = ws.cell(row=i, column=2)
        fiyat = s.get("fiyat")
        if fiyat is not None and fiyat == fiyat:  # NaN değil
            fc.value = round(float(fiyat), 2)
            fc.number_format = "#,##0.00"

        dc = ws.cell(row=i, column=3)
        deg = s.get("degisim")
        if deg is not None and deg == deg:
            deg = float(deg)
            dc.value = deg
            dc.number_format = '+0.00"%";-0.00"%";0.00"%"'
            renk = _YESIL if deg > 0 else (_KIRMIZI if deg < 0 else _KOYU)
            dc.font = Font(color=renk, bold=True)

        ws.cell(row=i, column=4, value=s.get("rating", "")).font = Font(color=_KOYU)
        ws.cell(row=i, column=5, value=s.get("sektor", "")).font = Font(color=_KOYU)

    for kol, gen in zip("ABCDE", (10, 12, 12, 16, 26)):
        ws.column_dimensions[kol].width = gen
    ws.freeze_panes = "A6"  # başlık sabit kalsın

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _deg_hucre(ws, satir, sutun, deger):
    c = ws.cell(row=satir, column=sutun)
    if deger is not None and deger == deger:
        deger = float(deger)
        c.value = deger
        c.number_format = '+0.00"%";-0.00"%";0.00"%"'
        c.font = Font(color=(_YESIL if deger > 0 else (_KIRMIZI if deger < 0 else _KOYU)), bold=True)
    return c


def toplu_rapor_excel(bolumler, tarama_zamani, veri_saati=None) -> bytes:
    """
    Çok kombinli rapor — SÜTUN düzeni (kullanıcının el yazısı listesi gibi).
    Her kombinasyon bir SÜTUN olur; altında o taramada çıkan firmalar alt alta.
    15-20 strateji yan yana, tek sayfada.
    bolumler: [{"ad": "Strateji adı", "satirlar": [{hisse, ...}]}]
    """
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tarama"

    n = max(len(bolumler), 1)
    # tüm hücrelere ızgara çizgisi (el yazısı tablo gibi)
    _gri = Side(style="thin", color="FF808080")
    kenar = Border(left=_gri, right=_gri, top=_gri, bottom=_gri)

    # üst bilgi (her strateji 2 sütun: hisse kodu + boş giriş kutusu)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 * n)
    ust = f"BIST Toplu Tarama · {tarama_zamani}"
    if veri_saati:
        ust += f"  ·  veri {veri_saati} (15 dk gecikmeli)"
    ws.cell(row=1, column=1, value=ust).font = Font(bold=True, size=11, color=_KOYU)

    max_firma = max((len(b.get("satirlar", [])) for b in bolumler), default=0)
    satir_sayisi = max(max_firma, 15)

    for j, bolum in enumerate(bolumler):
        c1 = 2 * j + 1   # hisse kodu sütunu
        c2 = 2 * j + 2   # boş giriş kutusu (kullanıcı veri yazar)
        firmalar = [s.get("hisse", "") for s in bolum.get("satirlar", [])]

        # strateji başlığı — iki sütuna yayılı, kenarlıklı
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        h = ws.cell(row=2, column=c1, value=str(bolum.get("ad", "")))
        h.font = Font(bold=True, color=_BEYAZ, size=10)
        h.fill = PatternFill("solid", fgColor=_BASLIK_BG)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=c1).border = kenar
        ws.cell(row=2, column=c2).border = kenar

        # satırlar: hisse kodu | boş kutucuk (veri girişi için)
        for i in range(satir_sayisi):
            hc = ws.cell(row=3 + i, column=c1)   # hisse kodu
            bc = ws.cell(row=3 + i, column=c2)   # boş kutu
            if i < len(firmalar):
                hc.value = firmalar[i]
                hc.font = Font(color=_KOYU)
            hc.alignment = Alignment(horizontal="center")
            hc.border = kenar
            bc.border = kenar
        ws.column_dimensions[get_column_letter(c1)].width = 10
        ws.column_dimensions[get_column_letter(c2)].width = 16

    ws.row_dimensions[2].height = 34
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    bolumler = [
        {"ad": "TİBO 7TM-7TZ-Basit", "satirlar": [{"hisse": h} for h in ["TEGYO", "MPARK", "TUPRS", "DURKN", "ANSGR"]]},
        {"ad": "MARGASİ", "satirlar": [{"hisse": h} for h in ["BEYAZ", "CRFSA", "KRSTL"]]},
        {"ad": "MARG SWING", "satirlar": [{"hisse": h} for h in ["ALARK", "DURKN", "KRPLS", "BQRAT", "FADE"]]},
        {"ad": "MACD 7 SMA", "satirlar": [{"hisse": h} for h in ["ASTOR", "KONT7"]]},
    ]
    data = toplu_rapor_excel(bolumler, "11.06.2026 18:45", "11.06 18:09")
    with open("/tmp/ornek_toplu.xlsx", "wb") as f:
        f.write(data)
    print("ornek_toplu.xlsx yazildi:", len(data), "bayt |", len(bolumler), "sütun")
