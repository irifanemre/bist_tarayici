"""
Performans Excel'i: dün çıkan kağıtlar + bugünkü durumları, en iyiden kötüye.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

_KOYU = "FF1F2328"
_GRI = "FF6E7681"
_YESIL = "FF1A7F37"
_KIRMIZI = "FFB42318"
_BASLIK_BG = "FF2F81F7"
_BEYAZ = "FFFFFFFF"
_gri = Side(style="thin", color="FF808080")
_kenar = Border(left=_gri, right=_gri, top=_gri, bottom=_gri)


def _basliklar(ws, satir, adlar, genislikler=None):
    from openpyxl.utils import get_column_letter
    for j, ad in enumerate(adlar, start=1):
        c = ws.cell(row=satir, column=j, value=ad)
        c.font = Font(bold=True, color=_BEYAZ, size=10)
        c.fill = PatternFill("solid", fgColor=_BASLIK_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _kenar
        if genislikler:
            ws.column_dimensions[get_column_letter(j)].width = genislikler[j - 1]


def _yuzde(ws, satir, sutun, deger):
    c = ws.cell(row=satir, column=sutun)
    c.border = _kenar
    c.alignment = Alignment(horizontal="center")
    if deger is None:
        c.value = "—"
        c.font = Font(color=_GRI)
        return
    c.value = float(deger)
    c.number_format = '+0.00"%";-0.00"%";0.00"%"'
    c.font = Font(bold=True, color=(_YESIL if deger > 0 else (_KIRMIZI if deger < 0 else _KOYU)))


def grid_performans_excel(bilgi, bolumler, karne) -> bytes:
    """Alışılmış 5x3 ızgara düzeni + her stratejide 'Güncel fiyat' ve 'Değişim' sütunu."""
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Performans"

    BLOK = 5          # yan yana kaç strateji
    SUT = 4           # her blok: hisse + fiyat + değişim + boşluk
    toplam = BLOK * SUT

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=toplam)
    ws.cell(row=1, column=1,
            value=f"DÜN ÇIKAN KAĞITLAR — GÜNCEL DURUM   ·   Tarama: {bilgi['tarama_zamani']}"
                  f"   →   Şimdi: {bilgi['simdi']}").font = Font(bold=True, size=12, color=_KOYU)

    if bilgi.get("uyari"):
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=toplam)
        u = ws.cell(row=2, column=1, value=bilgi["uyari"])
        u.font = Font(size=10, bold=True, color="FFB45309")

    max_firma = max((len(b["satirlar"]) for b in bolumler), default=0)
    satir_sayisi = max(max_firma, 6)

    for idx, bolum in enumerate(bolumler):
        band = idx // BLOK
        pos = idx % BLOK
        c1 = pos * SUT + 1                      # hisse
        c2, c3 = c1 + 1, c1 + 2                 # güncel fiyat, değişim
        hr = 4 + band * (satir_sayisi + 4)      # bu bandın başlık satırı

        # strateji adı (3 sütuna yayılı)
        ws.merge_cells(start_row=hr, start_column=c1, end_row=hr, end_column=c3)
        h = ws.cell(row=hr, column=c1, value=bolum["ad"])
        h.font = Font(bold=True, color=_BEYAZ, size=10)
        h.fill = PatternFill("solid", fgColor=_BASLIK_BG)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in (c1, c2, c3):
            ws.cell(row=hr, column=c).border = _kenar

        # alt başlıklar
        for c, ad in ((c1, "Hisse"), (c2, "Fiyat"), (c3, "Değ %")):
            b = ws.cell(row=hr + 1, column=c, value=ad)
            b.font = Font(bold=True, size=9, color=_KOYU)
            b.alignment = Alignment(horizontal="center")
            b.border = _kenar

        for i in range(satir_sayisi):
            r = hr + 2 + i
            s = bolum["satirlar"][i] if i < len(bolum["satirlar"]) else None
            hc = ws.cell(row=r, column=c1)
            fc = ws.cell(row=r, column=c2)
            if s:
                hc.value = s["hisse"]
                hc.font = Font(bold=True, color=_KOYU)
                if s.get("yeni") is not None:
                    fc.value = round(float(s["yeni"]), 2)
                    fc.number_format = "#,##0.00"
            hc.alignment = fc.alignment = Alignment(horizontal="center")
            hc.border = fc.border = _kenar
            _yuzde(ws, r, c3, s.get("degisim") if s else None)
            if not s:
                ws.cell(row=r, column=c3).value = None
                ws.cell(row=r, column=c3).border = _kenar

        ws.column_dimensions[get_column_letter(c1)].width = 10
        ws.column_dimensions[get_column_letter(c2)].width = 9
        ws.column_dimensions[get_column_letter(c3)].width = 9
        ws.column_dimensions[get_column_letter(c1 + 3)].width = 2
        ws.row_dimensions[hr].height = 30

    # --- strateji karnesi (en altta) ---
    bant = (len(bolumler) + BLOK - 1) // BLOK
    r = 4 + bant * (satir_sayisi + 4) + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="STRATEJİ KARNESİ (ortalama getiriye göre)"
            ).font = Font(bold=True, size=12, color=_KOYU)
    r += 1
    _basliklar(ws, r, ["Sıra", "Strateji", "Kağıt", "Kazanan", "Kaybeden", "Ortalama"])
    r += 1
    for i, k in enumerate(karne, start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = _kenar
        a = ws.cell(row=r, column=2, value=k["strateji"])
        a.font = Font(bold=True, color=_KOYU)
        a.border = _kenar
        for kol, val in ((3, k["adet"]), (4, k["kazanan"]), (5, k["kaybeden"])):
            c = ws.cell(row=r, column=kol, value=val)
            c.alignment = Alignment(horizontal="center")
            c.border = _kenar
        _yuzde(ws, r, 6, k.get("ortalama"))
        r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def performans_excel(bilgi, satirlar, karne) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Performans"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1,
            value=f"DÜN ÇIKAN KAĞITLAR — ERTESİ GÜN PERFORMANSI").font = Font(bold=True, size=13, color=_KOYU)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1,
            value=f"Tarama: {bilgi['tarama_zamani']}   →   Şimdi: {bilgi['simdi']}"
            ).font = Font(size=10, color=_GRI)

    # --- sıralı liste ---
    bas = 4
    _basliklar(ws, bas, ["Sıra", "Hisse", "Tarama fiyatı", "Güncel fiyat",
                         "Değişim", "Çıktığı strateji(ler)"],
               [7, 11, 14, 14, 12, 46])

    r = bas + 1
    for i, s in enumerate(satirlar, start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = _kenar
        h = ws.cell(row=r, column=2, value=s["hisse"])
        h.font = Font(bold=True, color=_KOYU)
        h.alignment = Alignment(horizontal="center")
        h.border = _kenar
        for kol, deg in ((3, s.get("eski")), (4, s.get("yeni"))):
            c = ws.cell(row=r, column=kol)
            c.border = _kenar
            c.alignment = Alignment(horizontal="center")
            if deg is not None:
                c.value = round(float(deg), 2)
                c.number_format = "#,##0.00"
            else:
                c.value = "—"
        _yuzde(ws, r, 5, s.get("degisim"))
        st = ws.cell(row=r, column=6, value=s.get("strateji", ""))
        st.font = Font(size=9, color=_GRI)
        st.border = _kenar
        r += 1

    # --- strateji karnesi ---
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="STRATEJİ KARNESİ (ortalama getiriye göre)"
            ).font = Font(bold=True, size=12, color=_KOYU)
    r += 1
    _basliklar(ws, r, ["Sıra", "Strateji", "Kağıt", "Kazanan", "Kaybeden", "Ortalama"])
    r += 1
    for i, k in enumerate(karne, start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = _kenar
        a = ws.cell(row=r, column=2, value=k["strateji"])
        a.font = Font(bold=True, color=_KOYU)
        a.border = _kenar
        for kol, val in ((3, k["adet"]), (4, k["kazanan"]), (5, k["kaybeden"])):
            c = ws.cell(row=r, column=kol, value=val)
            c.alignment = Alignment(horizontal="center")
            c.border = _kenar
        _yuzde(ws, r, 6, k.get("ortalama"))
        r += 1

    ws.freeze_panes = "A5"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
